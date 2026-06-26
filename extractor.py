from __future__ import annotations

import io
import re
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, Tuple, List

import fitz  # PyMuPDF
from PIL import Image
import pytesseract
from openpyxl import load_workbook

MONTHS = {
    "ENE": 1, "ENERO": 1, "JAN": 1,
    "FEB": 2, "FEBRERO": 2,
    "MAR": 3, "MARZO": 3,
    "ABR": 4, "ABRIL": 4, "APR": 4,
    "MAY": 5, "MAYO": 5,
    "JUN": 6, "JUNIO": 6,
    "JUL": 7, "JULIO": 7,
    "AGO": 8, "AGOSTO": 8, "AUG": 8,
    "SEP": 9, "SEPT": 9, "SEPTIEMBRE": 9,
    "OCT": 10, "OCTUBRE": 10,
    "NOV": 11, "NOVIEMBRE": 11,
    "DIC": 12, "DICIEMBRE": 12, "DEC": 12,
}

@dataclass
class ExtractedData:
    nombre: Optional[str] = None
    cedula: Optional[str] = None
    ingreso: Optional[str] = None          # YYYY-MM-DD
    valor_credito: Optional[float] = None
    seguro_vida_mensual: Optional[int] = None
    extraprima: Optional[str] = None
    fecha_nacimiento: Optional[str] = None # YYYY-MM-DD
    raw_confidence_notes: Optional[str] = None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text or " ").strip()


def _normalize_doc(raw: str | None) -> Optional[str]:
    if not raw:
        return None
    digits = re.sub(r"\D", "", raw).lstrip("0")
    if not digits:
        return None
    return f"{int(digits):,}".replace(",", ".")


def _parse_money(raw: str | None) -> Optional[float]:
    if not raw:
        return None
    try:
        return float(raw.replace("$", "").replace(",", "").strip())
    except ValueError:
        return None


def _parse_date(raw: str | None) -> Optional[str]:
    if not raw:
        return None
    s = raw.strip().upper().replace(".", "")
    # 05-06-2026 or 05/06/2026
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 1900 if y > 30 else 2000
        try:
            return datetime(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # 20-AGO-1981 or 05/JUN/2026
    m = re.search(r"(\d{1,2})[-/]([A-ZÁÉÍÓÚÑ]{3,10})[-/](\d{2,4})", s)
    if m:
        d = int(m.group(1))
        mon = m.group(2).replace("Á","A").replace("É","E").replace("Í","I").replace("Ó","O").replace("Ú","U")
        y = int(m.group(3))
        if y < 100:
            y += 1900 if y > 30 else 2000
        mo = MONTHS.get(mon[:3]) or MONTHS.get(mon)
        if mo:
            try:
                return datetime(y, mo, d).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def _ocr_page(page, scale: float = 1.8) -> str:
    """OCR a single PyMuPDF page. Raises EnvironmentError if tesseract missing."""
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
    img = Image.open(io.BytesIO(pix.tobytes("png")))
    try:
        return pytesseract.image_to_string(img, lang="spa+eng", config="--psm 6", timeout=30)
    except (RuntimeError, EnvironmentError, OSError) as e:
        if "tesseract" in str(e).lower():
            raise EnvironmentError(
                "Tesseract no está instalado. Instálalo con: brew install tesseract"
            ) from e
        return ""


# ── Core extraction from text ─────────────────────────────────────────────────

def _extract_core(embedded_norm: str, ocr_norm: str = "") -> Dict[str, Any]:
    """
    Extract all fields from already-normalized text strings.
    Used by both extract_fields() (single PDF) and extract_all_projections() (multi-PDF).
    """
    all_text = _normalize_spaces(embedded_norm + " " + ocr_norm)
    data = ExtractedData()

    # ── Nombre ──
    # Pattern: leading-zero cedula  date  (optional 0)  plazo  NOMBRE  linecode
    m = re.search(
        r"\b0{3,}\d+\s+\d{2}-\d{2}-\d{4}\s+0*\s*\d+\s+([A-ZÁÉÍÓÚÑ ]{8,90}?)\s+\d{4}\b",
        embedded_norm,
    )
    data.nombre = m.group(1).strip() if m else None

    # ── Cédula ──
    m = re.search(r"\b(0{3,}\d{6,12})\b", embedded_norm)
    if not m:
        m = re.search(r"(?:C[. ]?C|DOCUMENTO)[^\d]*(\d[\d .]{5,15})", all_text, re.I)
    data.cedula = _normalize_doc(m.group(1)) if m else None

    # ── Fecha ingreso ──
    m = re.search(
        r"(\d{2}-\d{2}-\d{4})\s+0*\s*\d+\s+[A-ZÁÉÍÓÚÑ ]{5,}\s+\d{4}",
        embedded_norm,
    )
    if not m:
        m = re.search(r"(\d{1,2}/[A-Z]{3}\.?/\d{4})", embedded_norm, re.I)
    data.ingreso = _parse_date(m.group(1)) if m else None

    # ── Valor crédito ──
    # Flexible: works for line codes 1008, 1011, etc.
    m = re.search(
        r"\d{4}\s+PMO\s+CONSUMO\s+SEG[^\d\n]*(\d{1,3}(?:,\d{3})+\.\d{2})",
        embedded_norm, re.I,
    )
    data.valor_credito = _parse_money(m.group(1)) if m else None

    # ── Seguro proporcional (regla de negocio clave) ──
    m = re.search(
        r"Seguro\s+(?:Cartera|Vida)\s+proporcional\s+(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)",
        embedded_norm, re.I,
    )
    if not m:
        m = re.search(
            r"(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s+Seguro\s+(?:Cartera|Vida)\s+proporcional",
            embedded_norm, re.I,
        )
    if m:
        data.seguro_vida_mensual = int(round(_parse_money(m.group(1)) or 0))
    else:
        m = re.search(
            r"\b1\s+20\d{4}\s+[\d,]+\s+0\s+[\d,]+\s+([\d,]+)\s+0\s+[\d,]+\s+[\d,]+",
            embedded_norm,
        )
        data.seguro_vida_mensual = int(m.group(1).replace(",", "")) if m else None

    # ── Fecha nacimiento (viene del OCR del formulario) ──
    fn = None
    pos = ocr_norm.upper().find("FECHA NACIMIENTO")
    if pos >= 0:
        window = ocr_norm[pos:pos + 220]
        mf = re.search(
            r"(\d{1,2}[-/][A-Z]{3}[-/]\d{2,4}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
            window, re.I,
        )
        if mf:
            fn = _parse_date(mf.group(1))
    if not fn:
        mf = re.search(
            r"\b[0-9. ]{7,15}\s+[|I]?\s*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\s+[|I]?",
            ocr_norm, re.I,
        )
        if mf:
            fn = _parse_date(mf.group(1))
    data.fecha_nacimiento = fn

    # ── Extraprima (de declaración de salud) ──
    data.extraprima = "EXTRAPRIMA INCLUIDA" if _detect_medical_yes_from_text(ocr_norm) else ""

    # ── Notas de confianza ──
    notes = []
    if not data.fecha_nacimiento:
        notes.append("Fecha nacimiento: revisar manualmente.")
    if not data.valor_credito:
        notes.append("Valor crédito no detectado.")
    data.raw_confidence_notes = " ".join(notes) or "OK"

    result = asdict(data)
    result["ingreso_excel"] = to_excel_date_display(data.ingreso)
    result["fecha_nacimiento_excel"] = to_excel_date_display(data.fecha_nacimiento)
    return result


# ── PDF text extraction ───────────────────────────────────────────────────────

def extract_pdf_text(pdf_path: str | Path, ocr_first_pages: int = 3) -> Tuple[str, str]:
    """Returns (embedded_text, ocr_text). OCR applied only to scanned/empty pages."""
    doc = fitz.open(str(pdf_path))
    embedded_parts, ocr_parts = [], []
    for idx, page in enumerate(doc):
        embedded = page.get_text("text") or ""
        embedded_parts.append(f"\n--- PAGE {idx+1} TEXT ---\n{embedded}")
        if idx < ocr_first_pages and len(embedded.strip()) < 30:
            ocr = _ocr_page(page)
            ocr_parts.append(f"\n--- PAGE {idx+1} OCR ---\n{ocr}")
    return "\n".join(embedded_parts), "\n".join(ocr_parts)


# ── Public API ────────────────────────────────────────────────────────────────

def extract_fields(pdf_path: str | Path) -> Dict[str, Any]:
    """Original single-PDF extractor (backward compatible)."""
    embedded, ocr = extract_pdf_text(pdf_path)
    result = _extract_core(_normalize_spaces(embedded), _normalize_spaces(ocr))
    # pixel-level extraprima check (needs the actual PDF file)
    if not result["extraprima"] and _detect_medical_yes_pixels(pdf_path):
        result["extraprima"] = "EXTRAPRIMA INCLUIDA"
    return result


def extract_all_projections(pdf_path: str | Path) -> List[Dict[str, Any]]:
    """
    Extract one record per PROYECCION DE CREDITO found in a multi-projection PDF.
    Returns list ordered as they appear in the document.
    """
    doc = fitz.open(str(pdf_path))
    pages_text = [page.get_text("text") or "" for page in doc]

    # Find where each new projection starts
    starts = [
        i for i, t in enumerate(pages_text)
        if "PROYECCION DE CREDITO" in t.upper()
    ]

    if not starts:
        # Fallback: treat entire PDF as one projection
        return [extract_fields(pdf_path)]

    results = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] if idx + 1 < len(starts) else len(pages_text)
        combined = "\n".join(pages_text[start:end])
        data = _extract_core(_normalize_spaces(combined), "")
        results.append(data)

    return results


def _fix_ocr_digits(s: str) -> str:
    """Replace common OCR letter-for-digit substitutions in a numeric string."""
    return (s.replace("B", "8").replace("b", "8")
             .replace("O", "0").replace("o", "0")
             .replace("I", "1").replace("l", "1")
             .replace("S", "5").replace("Z", "2")
             .replace("G", "6").replace("q", "9"))


def _extract_fn_from_cedula_text(text: str, upper: str) -> Optional[str]:
    """
    Extract fecha de nacimiento from a cedula photo page OCR result.
    Handles OCR artifacts:
    - Separators: -, /, ., space, S, or missing
    - Digits misread as letters: B→8, O→0, I→1, S→5
    - Date on same line without separator before year: "17-ABR1985"
    - Month abbreviation correct (ABR, FEB, etc.) or slightly off
    """
    def _try_parse_named_month(day_s: str, mon_s: str, year_s: str) -> Optional[str]:
        d = _fix_ocr_digits(day_s)
        y = _fix_ocr_digits(year_s)
        try:
            d_int = int(d)
            y_int = int(y) if len(y) == 4 else (int(y) + 1900 if int(y) > 30 else int(y) + 2000)
        except ValueError:
            return None
        if not (1 <= d_int <= 31 and 1940 <= y_int <= 2015):
            return None
        mon_clean = mon_s.upper()[:3]
        mo = MONTHS.get(mon_clean)
        if not mo:
            return None
        try:
            return datetime(y_int, mo, d_int).strftime("%Y-%m-%d")
        except ValueError:
            return None

    # Search near NACIMIENTO keyword (skip "LUGAR DE NACIMIENTO")
    for keyword in ("NACIMIENTO", "NACIM"):
        search_from = 0
        while True:
            idx = upper.find(keyword, search_from)
            if idx < 0:
                break
            context_before = upper[max(0, idx - 10):idx]
            if "LUGAR" not in context_before:
                window = text[max(0, idx - 60): idx + 250]

                # Named-month pattern: DD[-/. ]MMM[-/. ]YYYY
                # Year group allows letters so _fix_ocr_digits can fix e.g. 19B5→1985
                # NOTE: no numeric fallback here — cedulas always use named months
                # (e.g. 17-ABR-1985) and numeric fallback risks picking up the
                # FECHA DE EXPEDICION which is printed nearby.
                mf = re.search(
                    r"(\d{1,2})\s*[-/. S]?\s*([A-Za-z]{3,})\s*[-/. ]?\s*([0-9A-Za-z]{2,4})\b",
                    window, re.I,
                )
                if mf:
                    parsed = _try_parse_named_month(mf.group(1), mf.group(2), mf.group(3))
                    if parsed:
                        return parsed
                break
            search_from = idx + 1

    # Fallback: any named-month date on full page within plausible birth year range
    for mf in re.finditer(
        r"(\d{1,2})\s*[-/. S]?\s*([A-Za-z]{3,})\s*[-/. ]?\s*([0-9A-Za-z]{2,4})\b",
        text, re.I,
    ):
        parsed = _try_parse_named_month(mf.group(1), mf.group(2), mf.group(3))
        if parsed:
            return parsed

    return None


def _clean_ocr_date(raw: str) -> str:
    """
    Clean common OCR substitutions in a date string:
    S→5, B→8, O→0, I→1, Z→2, G→6 when adjacent to digits.
    Also removes stray spaces inside separators.
    """
    # Replace common character substitutions
    cleaned = raw
    replacements = [
        (r'(?<=\d)S(?=[\d\-/])', '5'),  # S after digit: 8S → 85
        (r'(?<=[\d\-/])S(?=\d)', '5'),  # S before digit
        (r'\bO(?=\d)', '0'),            # O at start of number
        (r'(?<=\d)O\b', '0'),
        (r'\bI(?=\d)', '1'),
        (r'(?<=\d)I\b', '1'),
    ]
    for pattern, repl in replacements:
        cleaned = re.sub(pattern, repl, cleaned)
    # Normalize separators: "13-04 - 85" → "13-04-85"
    cleaned = re.sub(r'\s*[-/]\s*', '-', cleaned)
    cleaned = re.sub(r'\s+', '-', cleaned.strip())
    return cleaned


def _extract_fn_from_declaration_text(text: str, upper: str) -> Optional[str]:
    """
    Extract fecha de nacimiento from a declaration (ASEGURABILIDAD) page.
    Strategies (in order):
    1. Near 'FECHA NACIMIENTO' / 'FECHA DE NACIMIENTO' keyword
    2. cedula_number + date on same line (e.g. '1099899969 06/06/1988')
       Handles OCR artifacts: spaces in separators, S→5, 2-digit years.
    """
    # Strategy 1: near keyword
    for kw in ("FECHA DE NACIMIENTO", "FECHA NACIMIENTO", "FECHANACIMIENTO"):
        pos = upper.find(kw)
        if pos >= 0:
            window = text[pos: pos + 300]
            mf = re.search(
                r"(\d{1,2}[-/][A-Z]{3}[-/]\d{2,4}|\d{2}[-/]\d{2}[-/]\d{4})",
                window, re.I,
            )
            if mf:
                parsed = _parse_date(mf.group(1))
                if parsed:
                    return parsed

    # Strategy 2: cedula number followed by date on the same line.
    # The form layout is: <cedula> <dd-mm-yy(yy)> <civil_status>
    # OCR may produce: "18402769 13-04 - 8S" → clean → "13-04-85" → 13/04/1985
    mf = re.search(
        r"\b\d{7,12}\s+"                          # cedula number
        r"(\d{2}\s*[-/ ]\s*\d{2}"                # dd-mm
        r"\s*[-/ ]+\s*\w{2,4})\b",               # separator + yy or yyyy (may have OCR noise)
        text,
    )
    if mf:
        raw_date = mf.group(1)
        cleaned = _clean_ocr_date(raw_date)
        # Extract only digits and dashes to try parsing
        parts = re.split(r'[-/]', cleaned)
        if len(parts) >= 3:
            try:
                d, mo, y_raw = parts[0], parts[1], parts[2]
                # Remove non-digit chars from year (OCR artifacts)
                y_digits = re.sub(r'\D', '5', y_raw)  # fallback S→5
                y_digits = re.sub(r'[^0-9]', '', y_digits)
                if y_digits:
                    parsed = _parse_date(f"{d}-{mo}-{y_digits}")
                    if parsed:
                        return parsed
            except Exception:
                pass

    # Strategy 3: any dd/mm/yy(yy) date with optional multi-char separators.
    # Also applies _fix_ocr_digits to the year field so "6S"→"65" and "BS"→"85".
    for mf in re.finditer(r"\b(\d{2})[-/ ]+(\d{2})[-/ ]+(\w{2,4})\b", text):
        d_str, m_str, y_raw = mf.group(1), mf.group(2), mf.group(3)
        y_str = _fix_ocr_digits(y_raw)
        if not y_str.isdigit():
            continue
        y = int(y_str)
        if y < 100:
            y += 1900 if y > 30 else 2000
        if 1940 <= y <= 2010:
            try:
                dt = datetime(y, int(m_str), int(d_str))
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue

    # Strategy 4: concatenated digits after FECHA NACIMIENTO keyword.
    # Also searches for bare NACIMIENTO when FECHA is misread by OCR (e.g. → RECNA).
    # Handles OCR that inserts noise digits into separators: "03/02/2000" → "0310212000"
    # (slashes read as "1"; noise) or "03022000" (no separators).
    for kw in ("FECHA NACIMIENTO", "FECHA DE NACIMIENTO", "FECHANACIMIENTO", "NACIMIENTO"):
        pos = upper.find(kw)
        if pos < 0:
            continue
        # For bare NACIMIENTO, skip "LUGAR DE NACIMIENTO" occurrences
        if kw == "NACIMIENTO":
            ctx = upper[max(0, pos - 12):pos]
            if "LUGAR" in ctx or "FECHA" in ctx:
                continue

        after_kw = text[pos + len(kw): pos + len(kw) + 80]
        digits = re.sub(r"\D", "", after_kw)

        # Find the year first, then extract DDMM from the 4 digits before it
        ym = re.search(r"(19|20)(\d{2})", digits)
        if ym:
            year = int(ym.group(0))
            before_year = digits[:ym.start()]
            # Strip any extra leading noise digits so we get exactly DDMM
            if len(before_year) >= 4:
                ddmm = before_year[-4:]  # last 4 digits before year = DDMM
                d_v, m_v = int(ddmm[:2]), int(ddmm[2:4])
                if 1 <= d_v <= 31 and 1 <= m_v <= 12 and 1940 <= year <= 2010:
                    try:
                        return datetime(year, m_v, d_v).strftime("%Y-%m-%d")
                    except ValueError:
                        pass

    return None


def _adaptive_threshold(img: "Image.Image") -> "Image.Image":
    """Simple adaptive binarization using local mean (no scipy needed)."""
    import numpy as np
    from PIL import ImageFilter
    arr = np.array(img).astype(float)
    blurred = np.array(img.filter(ImageFilter.BoxBlur(radius=20))).astype(float)
    binary = ((arr > blurred * 0.88) * 255).astype(np.uint8)
    return Image.fromarray(binary)


def _ocr_cedula_hires(page) -> Optional[str]:
    """
    High-resolution OCR pass for cedula back pages, focused on the FECHA NACIMIENTO row.
    Intentionally FAST: only 3 crop zones × 1 config = 3 OCR calls max.
    Returns named-month dates only (never numeric fallback) to avoid picking up
    expedition dates that appear on the same page.
    """
    from PIL import ImageEnhance, ImageFilter
    try:
        pix = page.get_pixmap(matrix=fitz.Matrix(3.5, 3.5), alpha=False)
        full = Image.open(io.BytesIO(pix.tobytes("png"))).convert("L")
    except Exception:
        return None

    w, h = full.size
    # Three targeted crop zones covering where NACIMIENTO row appears on Colombian cedulas
    for top_p, bot_p in [(0.30, 0.58), (0.38, 0.63), (0.22, 0.52)]:
        try:
            crop = full.crop((0, int(h * top_p), w, int(h * bot_p)))
            sharpened = crop.filter(ImageFilter.SHARPEN).filter(ImageFilter.SHARPEN)
            try:
                img = _adaptive_threshold(sharpened)
            except Exception:
                img = ImageEnhance.Contrast(sharpened).enhance(2.2)
            txt = pytesseract.image_to_string(
                img, lang="spa+eng", config="--psm 6 --oem 1 --dpi 350", timeout=18
            )
            t = _normalize_spaces(txt)
            result = _extract_fn_from_cedula_text(t, t.upper())
            if result:
                return result
        except Exception:
            continue
    return None


def _ocr_retry_fn(page, extractor_fn, scale: float = 2.0, contrast: float = 1.8) -> Optional[str]:
    """
    Re-OCR a page at higher scale+contrast, try multiple PSM/OEM combos.
    For cedula pages also tries cropping to the bottom half (cedula back)
    with adaptive thresholding.
    Returns the first non-None result from extractor_fn, or None.
    """
    from PIL import ImageEnhance
    try:
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        img_base = Image.open(io.BytesIO(pix.tobytes("png"))).convert("L")
        img_enh = ImageEnhance.Contrast(img_base).enhance(contrast)
        img_bin = _adaptive_threshold(img_base)
    except Exception:
        return None

    # Standard attempts: full page
    configs = [
        ("--psm 6  --oem 1 --dpi 300", img_enh),
        ("--psm 11 --oem 1 --dpi 300", img_enh),
        ("--psm 6  --oem 1 --dpi 300", img_bin),
        ("--psm 11 --oem 1 --dpi 300", img_bin),
        ("--psm 6  --oem 3 --dpi 300", img_enh),
    ]
    for cfg, img in configs:
        try:
            txt = pytesseract.image_to_string(img, lang="spa+eng", config=cfg, timeout=25)
            t = _normalize_spaces(txt)
            result = extractor_fn(t, t.upper())
            if result:
                return result
        except Exception:
            continue

    # Cedula-specific: try different vertical crop zones at 3x scale
    # (cedula back with birth date is typically in 40-60% vertical range)
    try:
        pix3 = page.get_pixmap(matrix=fitz.Matrix(3.0, 3.0), alpha=False)
        full = Image.open(io.BytesIO(pix3.tobytes("png"))).convert("L")
        w, h = full.size
        for top_p, bot_p in [(0.40, 0.60), (0.43, 0.58), (0.45, 0.62)]:
            crop = full.crop((0, int(h * top_p), w, int(h * bot_p)))
            crop_bin = _adaptive_threshold(crop)
            for cfg, img in [
                ("--psm 6 --oem 1 --dpi 300", crop_bin),
                ("--psm 11 --oem 1 --dpi 300", crop_bin),
            ]:
                try:
                    txt = pytesseract.image_to_string(img, lang="spa+eng", config=cfg, timeout=20)
                    t = _normalize_spaces(txt)
                    result = extractor_fn(t, t.upper())
                    if result:
                        return result
                except Exception:
                    continue
    except Exception:
        pass

    return None


def _ocr_text(page, scale: float = 1.8) -> str:
    """OCR a page; returns empty string on failure (does NOT raise)."""
    embedded = page.get_text("text") or ""
    if len(embedded.strip()) >= 30:
        return _normalize_spaces(embedded)
    try:
        return _normalize_spaces(_ocr_page(page, scale=scale))
    except EnvironmentError:
        raise
    except Exception:
        return ""


def extract_documents_ordered(pdf_path: str | Path) -> List[Dict[str, Any]]:
    """
    Extract an ORDERED list of document records from a multi-person documents PDF.

    Two-pass approach:
    Pass 1 – OCR every page at 1.8 scale and classify (declaration vs. other).
    Pass 2 – Process groups. When a declaration page is found:
       • Retroactively check the page immediately before it as a cedula photo
         (with high-res retry if 1.8 scale found nothing).
       • Then check the declaration page itself, plus any continuation pages.

    This avoids the bug where `in_person=True` causes cedula pages for
    subsequent persons to be misclassified as continuation pages.

    Returns: list ordered by person position for merging with extract_all_projections().
    Fields per entry: fecha_nacimiento, fecha_nacimiento_excel, extraprima.
    """
    doc = fitz.open(str(pdf_path))
    n = len(doc)

    # ── Pass 1: classify all pages ──────────────────────────────────────────
    page_texts: List[str] = []
    page_uppers: List[str] = []
    page_is_decl: List[bool] = []

    for page in doc:
        text = _ocr_text(page, scale=1.8)
        upper = text.upper()
        page_texts.append(text)
        page_uppers.append(upper)
        page_is_decl.append("ASEGURABILIDAD" in upper or "DECLARACION" in upper)

    # ── Pass 2: extract per person ──────────────────────────────────────────
    results: List[Dict] = []
    current: Dict = {}
    in_person = False

    for i in range(n):
        text = page_texts[i]
        upper = page_uppers[i]
        is_declaration = page_is_decl[i]

        if is_declaration:
            # Save previous person
            if in_person:
                results.append(_finalize_doc_record(current))

            # ── Try cedula photo page that precedes this declaration ──
            cedula_fn: Optional[str] = None
            if i > 0 and not page_is_decl[i - 1]:
                prev_page = doc[i - 1]
                prev_text = page_texts[i - 1]
                prev_upper = page_uppers[i - 1]
                # Quick first pass (already OCR'd at 1.8x)
                cedula_fn_initial = _extract_fn_from_cedula_text(prev_text, prev_upper)
                # Always run hires crop — it uses sharpening+binarization and often
                # corrects visual confusions (7→4, 8→6) that the 1.8x scan misses.
                cedula_fn_hires = _ocr_cedula_hires(prev_page)
                # Prefer hires result; fall back to initial only if hires found nothing
                cedula_fn = cedula_fn_hires or cedula_fn_initial

            # ── Start new person ──
            current = {"fn": cedula_fn, "extraprima": False}
            in_person = True

            # ── Try declaration page itself ──
            decl_fn_initial = _extract_fn_from_declaration_text(text, upper)
            # Always retry at 3.0x — higher scale reads ambiguous digits better
            # (e.g. "14" → "17", "6S" → more readable). Prefer hires result.
            decl_fn_hires = _ocr_retry_fn(doc[i], _extract_fn_from_declaration_text,
                                           scale=3.0, contrast=1.8)
            if not current["fn"]:
                current["fn"] = decl_fn_hires or decl_fn_initial

            current["extraprima"] = _detect_medical_yes_from_text(upper)

        elif in_person:
            # Continuation page (health questions, page 2+ of declaration)
            if not current.get("fn"):
                current["fn"] = _extract_fn_from_declaration_text(text, upper)
                if not current["fn"]:
                    current["fn"] = _ocr_retry_fn(doc[i], _extract_fn_from_declaration_text)
            current["extraprima"] = current["extraprima"] or _detect_medical_yes_from_text(upper)

        # Non-declaration pages before the first person are skipped —
        # they are picked up retroactively when the next declaration is found.

    # Flush last person
    if in_person:
        results.append(_finalize_doc_record(current))

    return results


def _finalize_doc_record(current: Dict) -> Dict[str, Any]:
    fn = current.get("fn")
    return {
        "fecha_nacimiento": fn,
        "fecha_nacimiento_excel": to_excel_date_display(fn),
        "extraprima": "EXTRAPRIMA INCLUIDA" if current.get("extraprima") else "",
    }


def extract_documents_by_cedula(pdf_path: str | Path) -> Dict[str, Dict[str, Any]]:
    """
    OCR a multi-page documents PDF and return a dict keyed by normalized cedula:
      { "9.728.672": { "fecha_nacimiento": "1981-08-20",
                       "fecha_nacimiento_excel": "08-20-81",
                       "extraprima": "" } }
    Pages without a recognizable cedula or fecha nacimiento are skipped.
    """
    doc = fitz.open(str(pdf_path))
    by_cedula: Dict[str, Dict] = {}

    # We process declaration pages (ASEGURABILIDAD) which hold the key fields.
    # Declaration spans ~2 pages; we accumulate text per person across consecutive pages.
    pending_cedula: Optional[str] = None
    pending_fn: Optional[str] = None
    pending_extraprima: bool = False

    for i, page in enumerate(doc):
        embedded = page.get_text("text") or ""
        if len(embedded.strip()) < 30:
            try:
                ocr_raw = _ocr_page(page, scale=1.8)
            except EnvironmentError:
                raise
            except Exception:
                ocr_raw = ""
        else:
            ocr_raw = embedded

        text = _normalize_spaces(ocr_raw)
        upper = text.upper()

        is_declaration = "ASEGURABILIDAD" in upper or "DECLARACION" in upper

        if is_declaration:
            # If we were accumulating a previous person and hit a new declaration, save them first
            if pending_cedula and pending_cedula not in by_cedula:
                by_cedula[pending_cedula] = {
                    "fecha_nacimiento": pending_fn,
                    "fecha_nacimiento_excel": to_excel_date_display(pending_fn),
                    "extraprima": "EXTRAPRIMA INCLUIDA" if pending_extraprima else "",
                }
            # Start new person
            pending_cedula = None
            pending_fn = None
            pending_extraprima = False

            # Extract cedula
            m = re.search(r"(?:C\.?\s*C\.?|CEDULA|DOCUMENTO)[^\d]*(\d[\d\s.]{5,14})", text, re.I)
            if not m:
                m = re.search(r"\b(\d{7,12})\b", text)
            if m:
                pending_cedula = _normalize_doc(m.group(1))

            # Extract fecha nacimiento
            pos = upper.find("FECHA NACIMIENTO")
            if pos >= 0:
                window = text[pos:pos + 250]
                mf = re.search(
                    r"(\d{1,2}[-/][A-Z]{3}[-/]\d{2,4}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
                    window, re.I,
                )
                if mf:
                    pending_fn = _parse_date(mf.group(1))

            # Health questions on same page
            pending_extraprima = _detect_medical_yes_from_text(upper)

        elif pending_cedula:
            # Continuation page for the same person (health questions page 2)
            if not pending_extraprima:
                pending_extraprima = _detect_medical_yes_from_text(upper)
            if not pending_fn:
                pos = upper.find("FECHA NACIMIENTO")
                if pos >= 0:
                    window = text[pos:pos + 250]
                    mf = re.search(
                        r"(\d{1,2}[-/][A-Z]{3}[-/]\d{2,4}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
                        window, re.I,
                    )
                    if mf:
                        pending_fn = _parse_date(mf.group(1))

    # Flush last pending person
    if pending_cedula and pending_cedula not in by_cedula:
        by_cedula[pending_cedula] = {
            "fecha_nacimiento": pending_fn,
            "fecha_nacimiento_excel": to_excel_date_display(pending_fn),
            "extraprima": "EXTRAPRIMA INCLUIDA" if pending_extraprima else "",
        }

    return by_cedula


# ── Medical detection ─────────────────────────────────────────────────────────

def _detect_medical_yes_from_text(text_upper: str) -> bool:
    """Keyword-only check for positive health declarations."""
    positive_words = [
        "ENFERMEDAD DEL CORAZON SI",
        "DIABETES SI",
        "CANCER SI",
        "PARALISIS SI",
        "DROGADICCION SI",
        "VIH/SIDA SI",
    ]
    return any(w in text_upper for w in positive_words)


def _detect_medical_yes_pixels(pdf_path: str | Path) -> bool:
    """Pixel-level heuristic for the Previsora insurance health declaration (page 2)."""
    try:
        doc = fitz.open(str(pdf_path))
        if len(doc) < 2:
            return False
        page = doc[1]
        pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
        img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("L")
        w, h = img.size
        bands = [(0.755, 0.785), (0.930, 0.960)]
        y1, y2 = int(h * 0.515), int(h * 0.705)
        for x1r, x2r in bands:
            x1, x2 = int(w * x1r), int(w * x2r)
            crop = img.crop((x1, y1, x2, y2))
            pixels = crop.load()
            dark = sum(
                1 for yy in range(crop.size[1]) for xx in range(crop.size[0])
                if pixels[xx, yy] < 80
            )
            total = crop.size[0] * crop.size[1]
            if total and dark / total > 0.09:
                return True
    except Exception:
        return False
    return False


def _detect_medical_yes(pdf_path: str | Path, ocr_norm: str = "") -> bool:
    """Legacy combined check (kept for backward compat)."""
    return _detect_medical_yes_from_text(ocr_norm.upper()) or _detect_medical_yes_pixels(pdf_path)


# ── Excel output ──────────────────────────────────────────────────────────────

def fill_excel_template(
    template_path: str | Path,
    output_path: str | Path,
    data: Dict[str, Any],
    month_sheet: str = "JUNIO",
) -> str:
    from copy import copy

    wb = load_workbook(template_path)
    ws = wb[month_sheet]
    row = 4
    while ws.cell(row=row, column=1).value:
        row += 1

    template_row = row - 1 if row > 4 else 4
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=template_row, column=col)
        dst = ws.cell(row=row, column=col)
        if src.has_style:
            dst._style = copy(src._style)
        for attr in ("number_format", "alignment", "border", "fill", "font"):
            val = getattr(src, attr, None)
            if val:
                setattr(dst, attr, copy(val))

    ws.cell(row=row, column=1).value = data.get("nombre")
    ws.cell(row=row, column=4).value = data.get("cedula")
    ws.cell(row=row, column=5).value = _parse_date_to_datetime(data.get("ingreso"))
    ws.cell(row=row, column=6).value = data.get("valor_credito")
    ws.cell(row=row, column=7).value = data.get("seguro_vida_mensual")
    ws.cell(row=row, column=8).value = data.get("extraprima")
    ws.cell(row=row, column=9).value = _parse_date_to_datetime(data.get("fecha_nacimiento"))
    ws.cell(row=row, column=5).number_format = "dd/mm/yyyy"
    ws.cell(row=row, column=9).number_format = "dd/mm/yyyy"
    wb.save(output_path)
    return str(output_path)


def _parse_date_to_datetime(value: Optional[str]):
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d")


def to_excel_date_display(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return ""


if __name__ == "__main__":
    import argparse, json
    parser = argparse.ArgumentParser()
    parser.add_argument("pdf")
    parser.add_argument("--template")
    parser.add_argument("--out", default="salida.xlsx")
    parser.add_argument("--month", default="JUNIO")
    parser.add_argument("--mode", choices=["single", "projections", "documents"], default="single")
    args = parser.parse_args()

    if args.mode == "projections":
        results = extract_all_projections(args.pdf)
        print(json.dumps(results, ensure_ascii=False, indent=2))
    elif args.mode == "documents":
        results = extract_documents_by_cedula(args.pdf)
        print(json.dumps(results, ensure_ascii=False, indent=2))
    else:
        result = extract_fields(args.pdf)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        if args.template:
            fill_excel_template(args.template, args.out, result, args.month)
            print(f"Excel generado: {args.out}")
